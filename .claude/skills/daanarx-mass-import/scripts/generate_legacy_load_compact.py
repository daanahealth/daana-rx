#!/usr/bin/env python3
"""
Compact single-statement legacy load: emits a CTE that stages the aggregated
rows inline (no UUIDs emitted) and inserts locations/lots/drugs/units using
gen_random_uuid() + natural-key joins. ~5x smaller than the explicit-UUID form,
so it can be executed in one Supabase MCP call.

Same conventions/aggregation as generate_legacy_load.py.

Usage:
  python3 generate_legacy_load_compact.py INPUT.xlsx --clinic <uuid> --user <uuid> \
      --source-tag "XLSX_IMPORT 26.06.14" --outdir OUT
"""
import argparse, datetime, os, re, collections

SHEET_FORM = {"DATA-CARDS": "Card", "DATA-BOTTLES": "Bottle", "DATA-PSYCH": "Psych"}
SPEC_FIX = {"OCT": "OTC", "PSCYH": "PSYCH", "URO": "UROLOGY"}
NO_EXPIRY = "2099-12-31"


def clean(c):
    return "" if c is None else str(c).strip()


def norm_strength(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return f"{v.month}-{v.day}"
    s = clean(v)
    return s[:-2] if re.fullmatch(r"\d+\.0", s) else s


def numeric_strength(dose):
    m = re.search(r"\d+(\.\d+)?", dose or "")
    return m.group(0) if m else "0"


def canon_specialty(label):
    s = re.sub(r"\s+", " ", clean(label)).upper()
    return " ".join(SPEC_FIX.get(p, p) for p in s.split(" ")) if s else ""


def title_med(name):
    return " ".join(w.capitalize() if (w.isupper() or w.islower()) else w
                    for w in clean(name).split())


def q(s):
    return "'" + str(s).replace("'", "''") + "'"


def read_rows(path, form_arg):
    rows = []
    if path.lower().endswith((".xlsx", ".xlsm")):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        for sheet, form in SHEET_FORM.items():
            if sheet in wb.sheetnames:
                for r in wb[sheet].iter_rows(values_only=True):
                    rows.append((form, list(r) + [None] * 6))
    else:
        import csv
        with open(path, newline="") as f:
            for r in csv.reader(f):
                rows.append((form_arg, list(r) + [None] * 6))
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("input")
    ap.add_argument("--clinic", required=True)
    ap.add_argument("--user", required=True)
    ap.add_argument("--form", default="Card")
    ap.add_argument("--source-tag", default="XLSX_IMPORT")
    ap.add_argument("--outdir", default=".")
    args = ap.parse_args()
    os.makedirs(args.outdir, exist_ok=True)
    fname = os.path.basename(args.input)

    groups = collections.OrderedDict()
    lot_spec = {}
    skipped = 0
    for form, row in read_rows(args.input, args.form):
        lot, spec, med, st, un, qid = [clean(x) for x in row[:5]] + [clean(row[5])]
        dose = norm_strength(row[3])
        med = re.sub(r"\s+", " ", med).strip()
        if not med or not qid:
            skipped += 1
            continue
        lot = lot.upper()
        key = (lot, med.upper(), dose, un.upper(), form)
        g = groups.setdefault(key, dict(lot=lot, med=med, dose=dose, unit=un,
                                        form=form, qty=0))
        g["qty"] += 1
        lot_spec.setdefault(lot, canon_specialty(spec) or lot)

    # staged rows: lot, specialty, drug_name, strength(num), su, form, qty
    staged = []
    for g in groups.values():
        name = title_med(g["med"])
        if g["dose"] and g["dose"].upper() != "X":
            name += f" {g['dose']}"
        if g["unit"] and g["unit"].upper() != "X":
            name += f" {g['unit'].lower()}"
        su = "" if g["unit"].upper() == "X" else g["unit"].lower()
        # drugs.strength_unit is NOT NULL -> use '' (never NULL) for unit-less items
        su_sql = q(su)
        staged.append((g["lot"], lot_spec[g["lot"]], name,
                       numeric_strength(g["dose"]), su_sql, g["form"], g["qty"]))

    C, U, TAG = q(args.clinic), q(args.user), q(args.source_tag)
    lines = []
    for i, (lot, spec, name, strv, su_sql, form, qty) in enumerate(staged):
        cast = "::numeric" if i == 0 else ""
        lines.append(f"  ({q(lot)}, {q(spec)}, {q(name)}, {strv}{cast}, {su_sql}, {q(form)}, {qty})")
    values = ",\n".join(lines)

    sql = f"""BEGIN;
WITH s(lot, specialty, dname, strength, su, form, qty) AS (VALUES
{values}
),
loc AS (
  INSERT INTO locations (location_id, name, temp, clinic_id)
  SELECT gen_random_uuid(), d.specialty || ' - ' || d.lot, 'room temp', {C}::uuid
  FROM (SELECT DISTINCT lot, specialty FROM s) d
  RETURNING location_id, name
),
lt AS (
  INSERT INTO lots (lot_id, source, note, location_id, clinic_id, max_capacity)
  SELECT gen_random_uuid(), {TAG}, 'LocationCode: ' || d.lot, loc.location_id, {C}::uuid, 100
  FROM (SELECT DISTINCT lot, specialty FROM s) d
  JOIN loc ON loc.name = d.specialty || ' - ' || d.lot
  RETURNING lot_id, note
),
dr AS (
  -- drugs.ndc_id is NOT NULL + UNIQUE; '' is taken. Match the prior import's
  -- synthetic 'PRELIM-<16hex>' scheme with a generated unique value per drug.
  INSERT INTO drugs (drug_id, medication_name, generic_name, strength, strength_unit, ndc_id, form)
  SELECT gen_random_uuid(), d.dname, d.dname, d.strength, d.su,
         'PRELIM-' || substr(replace(gen_random_uuid()::text, '-', ''), 1, 16), d.form
  FROM (SELECT DISTINCT dname, strength, su, form FROM s) d
  RETURNING drug_id, medication_name, strength, strength_unit, form
)
INSERT INTO units (unit_id, total_quantity, available_quantity, lot_id, expiry_date,
                   user_id, drug_id, qr_code, optional_notes, clinic_id)
SELECT gen_random_uuid(), s.qty, s.qty, lt.lot_id, '{NO_EXPIRY}'::date, {U}::uuid,
       dr.drug_id, NULL, 'Imported from {fname} (' || s.form || '(s))', {C}::uuid
FROM s
JOIN lt ON lt.note = 'LocationCode: ' || s.lot
JOIN dr ON dr.medication_name = s.dname AND dr.form = s.form
       AND dr.strength = s.strength AND dr.strength_unit IS NOT DISTINCT FROM s.su;
COMMIT;

SELECT (SELECT count(*) FROM units WHERE clinic_id={C}) units,
       (SELECT sum(total_quantity) FROM units WHERE clinic_id={C}) total_qty,
       (SELECT count(*) FROM lots WHERE clinic_id={C}) lots,
       (SELECT count(*) FROM locations WHERE clinic_id={C}) locations,
       (SELECT count(DISTINCT drug_id) FROM units WHERE clinic_id={C}) distinct_drugs;
"""
    path = os.path.join(args.outdir, "03_load_compact.sql")
    with open(path, "w") as f:
        f.write(sql)
    print(f"rows staged={len(staged)} total_qty={sum(x[6] for x in staged)} "
          f"distinct_lots={len(lot_spec)} skipped={skipped}")
    print(f"bytes={len(sql)}  wrote {path}")


if __name__ == "__main__":
    main()
