#!/usr/bin/env python3
"""
DaanaRX MASS inventory parser/normalizer.

Reads a clinic data export (.xlsx with DATA-* type sheets + SPECIALTY LOCATION,
or a single .csv) and produces records in the DaanaRX `items` core-schema shape,
aggregated to one record per (form, specialty, medication, strength, unit) with
quantity = count of repeating physical containers.

Outputs (into --outdir):
  - mass_items_load.json   the aggregated records ready to insert
  - locations.csv          derived location code/specialty/capacity table
  - review-needed.csv       ambiguous rows a human should eyeball before load
  - summary.txt            human-readable counts

Column order in every data sheet/csv (no header):
  lot_code, specialty_label, medication, strength, strength_unit, source_id(8-digit)

Usage:
  python3 parse_inventory.py INPUT.xlsx --outdir OUT
  python3 parse_inventory.py INPUT.csv  --form Card --outdir OUT
"""
import argparse, csv, datetime, json, os, re, sys, collections

# sheet name -> form/type. Extend here if the clinic adds new type tabs.
SHEET_FORM = {
    "DATA-CARDS": "Card",
    "DATA-BOTTLES": "Bottle",
    "DATA-PSYCH": "Psych",
}

# light, conservative specialty-label canonicalization (typos seen in the wild)
SPEC_FIX = {
    "OCT": "OTC", "PSCYH": "PSYCH", "URO": "UROLOGY",
}


def clean(c):
    return "" if c is None else str(c).strip()


def norm_strength(v):
    """Return (text_value, recovered_flag). Dosage is kept as TEXT so combos
    like '8.6-50' and '100/1000' survive. Excel sometimes coerces combo doses
    (e.g. '10-15') into real dates -> recover as 'month-day'."""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return f"{v.month}-{v.day}", True
    s = clean(v)
    if re.fullmatch(r"\d+\.0", s):  # 325.0 -> 325
        s = s[:-2]
    return s, False


def canon_specialty(label):
    s = re.sub(r"\s+", " ", clean(label)).upper()
    if not s:
        return ""
    parts = s.split(" ")
    parts = [SPEC_FIX.get(p, p) for p in parts]
    s = " ".join(parts)
    # "URO 1" -> "UROLOGY 1" handled by token fix above
    return s


def lot_base(lot):
    """Leading base that keys SPECIALTY LOCATION (a letter A-Z, or a number)."""
    s = clean(lot).upper()
    s = re.sub(r"^LOT\s*", "", s)
    if not s:
        return ""
    if s[0].isdigit():
        m = re.match(r"\d+", s)
        return m.group(0)
    return s[0]


def loc_code(specialty):
    """Deterministic short UNIQUE-ish code for a specialty (no spaces)."""
    s = re.sub(r"[^A-Z0-9]+", "", specialty.upper())
    return s or "MISC"


def read_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    specmap = {}
    if "SPECIALTY LOCATION" in wb.sheetnames:
        for r in wb["SPECIALTY LOCATION"].iter_rows(min_row=2, values_only=True):
            k = clean(r[0]).upper()
            if k.endswith(".0"):
                k = k[:-2]
            if k:
                specmap[k] = canon_specialty(r[1]) if len(r) > 1 else ""
    rows = []
    for sheet, form in SHEET_FORM.items():
        if sheet not in wb.sheetnames:
            continue
        for row in wb[sheet].iter_rows(values_only=True):
            rows.append((form, list(row) + [None] * (6 - len(row))))
    return rows, specmap


def read_csv(path, form):
    rows = []
    with open(path, newline="") as f:
        for r in csv.reader(f):
            rows.append((form, (r + [None] * 6)[:6]))
    return rows, {}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("input")
    ap.add_argument("--form", default="Card", help="form/type for a bare .csv input")
    ap.add_argument("--outdir", default=".")
    args = ap.parse_args()
    os.makedirs(args.outdir, exist_ok=True)

    if args.input.lower().endswith((".xlsx", ".xlsm")):
        raw, specmap = read_xlsx(args.input)
    else:
        raw, specmap = read_csv(args.input, args.form)

    units = []        # one dict per valid physical row
    skipped = 0
    for form, row in raw:
        lot, spec, med, st, un, qid = [clean(x) for x in row[:5]] + [clean(row[5])]
        strength, recovered = norm_strength(row[3])
        med = re.sub(r"\s+", " ", med).strip()
        un = un.upper()
        if not med or not qid:        # blank/junk row
            skipped += 1
            continue
        cspec = canon_specialty(spec)
        base = lot_base(lot)
        map_spec = specmap.get(base, "")
        units.append(dict(form=form, lot=lot.upper(), base=base, spec=cspec,
                          map_spec=map_spec, med=med, strength=strength,
                          unit=un, qid=qid, recovered=recovered))

    # ---- aggregate: quantity = count of repeating (med, strength, unit) within (form, specialty)
    groups = collections.OrderedDict()
    for u in units:
        key = (u["form"], u["spec"], u["med"].upper(), u["strength"], u["unit"])
        g = groups.setdefault(key, dict(form=u["form"], specialty=u["spec"],
                                        medication=u["med"], strength=u["strength"],
                                        unit=u["unit"], quantity=0, lot_codes=set(),
                                        source_ids=[], recovered=False))
        g["quantity"] += 1
        g["lot_codes"].add(u["lot"])
        g["source_ids"].append(u["qid"])
        g["recovered"] = g["recovered"] or u["recovered"]

    # ---- locations + per-location DRX counter -> unit_code
    specialties = sorted({g["specialty"] for g in groups.values() if g["specialty"]})
    codes = {}
    used = set()
    for sp in specialties:
        c = loc_code(sp)
        base_c = c
        i = 2
        while c in used:                  # guarantee uniqueness
            c = f"{base_c}{i}"; i += 1
        used.add(c)
        codes[sp] = c

    counters = collections.Counter()
    records = []
    for g in groups.values():
        sp = g["specialty"] or "MISC"
        code = codes.get(sp, loc_code(sp))
        counters[code] += 1
        unit_code = f"DRX-MASS-{code}-{counters[code]:05d}"
        is_num = bool(re.fullmatch(r"\d+(\.\d+)?", g["strength"]))
        records.append({
            "unit_code": unit_code,
            "status": "active",
            "expiry_date": None,
            "location_code": code,
            "specialty": sp,
            "attributes": {
                "medication_name": g["medication"],
                "dosage": g["strength"],
                "unit": g["unit"],
                "form": g["form"],
                "quantity": g["quantity"],
                "lot_codes": sorted(g["lot_codes"]),
                "source_ids": g["source_ids"],
                "source_file": os.path.basename(args.input),
            },
            "_needs_review": (not is_num) or g["recovered"] or not g["unit"] or not g["specialty"],
        })

    # ---- write outputs
    out = args.outdir
    with open(os.path.join(out, "mass_items_load.json"), "w") as f:
        json.dump({"item_type": "medication", "records": records}, f, indent=2)

    with open(os.path.join(out, "locations.csv"), "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["location_code", "specialty", "capacity"])
        for sp in specialties:
            w.writerow([codes[sp], sp, 50])

    review = [r for r in records if r["_needs_review"]]
    with open(os.path.join(out, "review-needed.csv"), "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["unit_code", "specialty", "medication", "dosage", "unit",
                    "form", "quantity", "lot_codes", "reason"])
        for r in review:
            a = r["attributes"]
            reasons = []
            if not re.fullmatch(r"\d+(\.\d+)?", a["dosage"]):
                reasons.append("non-numeric dosage")
            if not a["unit"]:
                reasons.append("missing unit")
            if not r["specialty"]:
                reasons.append("missing specialty")
            w.writerow([r["unit_code"], r["specialty"], a["medication_name"],
                        a["dosage"], a["unit"], a["form"], a["quantity"],
                        "|".join(a["lot_codes"]), "; ".join(reasons) or "recovered-from-date"])

    # label vs authoritative-map mismatches (informational)
    mism = sorted({(u["lot"], u["spec"], u["map_spec"]) for u in units
                   if u["map_spec"] and u["spec"] and
                   u["spec"].replace(" ", "") not in u["map_spec"].replace(" ", "")})

    by_form = collections.Counter(u["form"] for u in units)
    with open(os.path.join(out, "summary.txt"), "w") as f:
        def p(*a): print(*a); print(*a, file=f)
        p(f"input: {args.input}")
        p(f"physical units (rows): {len(units)}   skipped junk rows: {skipped}")
        p(f"  by form: {dict(by_form)}")
        p(f"aggregated item records: {len(records)}   (quantity = count of repeats)")
        p(f"total quantity across records: {sum(r['attributes']['quantity'] for r in records)}")
        p(f"locations/specialties: {len(specialties)}")
        p(f"records needing review: {len(review)}")
        p(f"lot-label vs SPECIALTY-LOCATION-map mismatches: {len(mism)}")
        for m in mism:
            p(f"    lot={m[0]:6} data_label={m[1]!r:30} map={m[2]!r}")
    print(f"\nwrote outputs to {out}/")


if __name__ == "__main__":
    main()
