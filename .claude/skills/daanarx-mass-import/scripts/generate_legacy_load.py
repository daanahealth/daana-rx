#!/usr/bin/env python3
"""
Generate legacy-schema (drugs/lots/locations/units) load SQL for a MASS clinic,
matching the convention established by the prior CSV_IMPORT seed:

  drugs:     medication_name = "<Title Name> <dose> <unit>", strength numeric,
             strength_unit lowercased, form = sheet type (Card/Bottle/Psych)
  locations: one per lot code, name = "<specialty> - <lot>", temp "room temp"
  lots:      one per lot code, source = SOURCE_TAG, note = "LocationCode: <lot>"
  units:     total_quantity = available_quantity = count of repeats PER LOT,
             expiry_date = 2099-12-31 (no-expiry placeholder, column is NOT NULL),
             qr_code = NULL, optional_notes = "Imported from <file> (<form>)"

Aggregation key = (lot, medication, strength, unit, form)  -> one unit row.

Pre-generates UUIDs in Python so FKs link deterministically. Emits:
  01_inspect_legacy.sql, 02_delete_legacy.sql, 03_load_legacy.sql

Usage:
  python3 generate_legacy_load.py INPUT.xlsx --clinic <uuid> --user <uuid> --outdir OUT
"""
import argparse, datetime, json, os, re, uuid, collections

SHEET_FORM = {"DATA-CARDS": "Card", "DATA-BOTTLES": "Bottle", "DATA-PSYCH": "Psych"}
SPEC_FIX = {"OCT": "OTC", "PSCYH": "PSYCH", "URO": "UROLOGY"}
NO_EXPIRY = "2099-12-31"


def clean(c):
    return "" if c is None else str(c).strip()


def norm_strength(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return f"{v.month}-{v.day}"
    s = clean(v)
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return s


def numeric_strength(dose):
    """Best-effort numeric for the NOT NULL drugs.strength column."""
    m = re.search(r"\d+(\.\d+)?", dose or "")
    return m.group(0) if m else "0"


def canon_specialty(label):
    s = re.sub(r"\s+", " ", clean(label)).upper()
    return " ".join(SPEC_FIX.get(p, p) for p in s.split(" ")) if s else ""


def title_med(name):
    return " ".join(w.capitalize() if w.isupper() or w.islower() else w
                    for w in clean(name).split())


def sql_str(s):
    return "'" + str(s).replace("'", "''") + "'"


def sql_val(s):
    return "NULL" if s is None or s == "" else sql_str(s)


def read_rows(path, form_arg):
    rows = []
    if path.lower().endswith((".xlsx", ".xlsm")):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        for sheet, form in SHEET_FORM.items():
            if sheet in wb.sheetnames:
                for r in wb[sheet].iter_rows(values_only=True):
                    rows.append((form, list(r) + [None] * 6))
    else:
        import csv
        with open(path, newline="") as f:
            for r in csv.reader(f):
                rows.append((form_arg, (list(r) + [None] * 6)))
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("input")
    ap.add_argument("--clinic", required=True)
    ap.add_argument("--user", required=True)
    ap.add_argument("--form", default="Card")
    ap.add_argument("--source-tag", default="XLSX_IMPORT")
    ap.add_argument("--outdir", default=".")
    args = ap.parse_args()
    os.makedirs(args.outdir, exist_ok=True)

    raw = read_rows(args.input, args.form)
    fname = os.path.basename(args.input)

    # aggregate per (lot, med, strength, unit, form)
    groups = collections.OrderedDict()
    lot_spec = {}
    skipped = 0
    for form, row in raw:
        lot, spec, med, st, un, qid = [clean(x) for x in row[:5]] + [clean(row[5])]
        dose = norm_strength(row[3])
        med = re.sub(r"\s+", " ", med).strip()
        if not med or not qid:
            skipped += 1
            continue
        lot = lot.upper()
        key = (lot, med.upper(), dose, un.upper(), form)
        g = groups.setdefault(key, dict(lot=lot, med=med, dose=dose, unit=un,
                                        form=form, qty=0, ids=[]))
        g["qty"] += 1
        g["ids"].append(qid)
        lot_spec.setdefault(lot, canon_specialty(spec) or lot)

    # ids: one location + lot per lot code; one drug per (med,dose,unit,form)
    loc_id = {lot: str(uuid.uuid4()) for lot in lot_spec}
    lot_id = {lot: str(uuid.uuid4()) for lot in lot_spec}
    drug_key = {}
    drug_rows = []
    for g in groups.values():
        dk = (g["med"].upper(), g["dose"], g["unit"].upper(), g["form"])
        if dk not in drug_key:
            did = str(uuid.uuid4())
            drug_key[dk] = did
            name = title_med(g["med"])
            if g["dose"] and g["dose"].upper() != "X":
                name += f" {g['dose']}"
            if g["unit"] and g["unit"].upper() != "X":
                name += f" {g['unit'].lower()}"
            un_clean = "" if g["unit"].upper() == "X" else g["unit"].lower()
            drug_rows.append((did, name, numeric_strength(g["dose"]),
                              un_clean, g["form"]))

    C, U = args.clinic, args.user

    with open(os.path.join(args.outdir, "01_inspect_legacy.sql"), "w") as f:
        f.write(
            f"-- Current MASS clinic ({C}) legacy inventory before changes\n"
            f"SELECT 'units' t, count(*) FROM units WHERE clinic_id={sql_str(C)}\n"
            f"UNION ALL SELECT 'lots', count(*) FROM lots WHERE clinic_id={sql_str(C)}\n"
            f"UNION ALL SELECT 'locations', count(*) FROM locations WHERE clinic_id={sql_str(C)}\n"
            f"UNION ALL SELECT 'transactions', count(*) FROM transactions WHERE clinic_id={sql_str(C)};\n"
        )

    with open(os.path.join(args.outdir, "02_delete_legacy.sql"), "w") as f:
        f.write(
            f"-- HARD DELETE existing MASS clinic inventory (clinic {C}).\n"
            f"-- drugs are global/shared across clinics -> NOT deleted.\n"
            f"BEGIN;\n"
            f"DELETE FROM transactions WHERE clinic_id={sql_str(C)};\n"
            f"DELETE FROM units WHERE clinic_id={sql_str(C)};\n"
            f"DELETE FROM lots  WHERE clinic_id={sql_str(C)};\n"
            f"DELETE FROM locations WHERE clinic_id={sql_str(C)};\n"
            f"COMMIT;\n"
        )

    with open(os.path.join(args.outdir, "03_load_legacy.sql"), "w") as f:
        f.write(f"-- Load {fname} into MASS clinic {C} (importer {U}).\n")
        f.write(f"-- {len(loc_id)} locations, {len(lot_id)} lots, "
                f"{len(drug_rows)} drugs, {len(groups)} units.\nBEGIN;\n\n")

        f.write("INSERT INTO locations (location_id, name, temp, clinic_id) VALUES\n")
        rows = [f"  ({sql_str(loc_id[lot])}, {sql_str(f'{lot_spec[lot]} - {lot}')}, "
                f"'room temp', {sql_str(C)})" for lot in lot_spec]
        f.write(",\n".join(rows) + ";\n\n")

        f.write("INSERT INTO lots (lot_id, source, note, location_id, clinic_id, max_capacity) VALUES\n")
        rows = [f"  ({sql_str(lot_id[lot])}, {sql_str(args.source_tag)}, "
                f"{sql_str('LocationCode: ' + lot)}, {sql_str(loc_id[lot])}, {sql_str(C)}, 100)"
                for lot in lot_spec]
        f.write(",\n".join(rows) + ";\n\n")

        f.write("INSERT INTO drugs (drug_id, medication_name, generic_name, strength, strength_unit, form) VALUES\n")
        rows = [f"  ({sql_str(d)}, {sql_str(nm)}, {sql_str(nm)}, {st}, {sql_val(un)}, {sql_str(fm)})"
                for (d, nm, st, un, fm) in drug_rows]
        f.write(",\n".join(rows) + ";\n\n")

        f.write("INSERT INTO units (unit_id, total_quantity, available_quantity, lot_id, "
                "expiry_date, user_id, drug_id, qr_code, optional_notes, clinic_id) VALUES\n")
        rows = []
        for g in groups.values():
            dk = (g["med"].upper(), g["dose"], g["unit"].upper(), g["form"])
            note = f"Imported from {fname} ({g['form']}(s))"
            rows.append(
                f"  ({sql_str(str(uuid.uuid4()))}, {g['qty']}, {g['qty']}, {sql_str(lot_id[g['lot']])}, "
                f"'{NO_EXPIRY}', {sql_str(U)}, {sql_str(drug_key[dk])}, NULL, {sql_str(note)}, {sql_str(C)})")
        f.write(",\n".join(rows) + ";\n\n")
        f.write("COMMIT;\n\n")
        f.write(
            f"-- Verify\nSELECT (SELECT count(*) FROM units WHERE clinic_id={sql_str(C)}) units,\n"
            f"       (SELECT sum(total_quantity) FROM units WHERE clinic_id={sql_str(C)}) total_qty,\n"
            f"       (SELECT count(*) FROM lots WHERE clinic_id={sql_str(C)}) lots,\n"
            f"       (SELECT count(*) FROM locations WHERE clinic_id={sql_str(C)}) locations;\n")

    # sidecar traceability: which physical 8-digit IDs back each unit row
    import csv as _csv
    with open(os.path.join(args.outdir, "source_ids_map.csv"), "w", newline="") as f:
        w = _csv.writer(f)
        w.writerow(["lot", "medication", "dose", "unit", "form", "quantity", "source_ids"])
        for g in groups.values():
            w.writerow([g["lot"], g["med"], g["dose"], g["unit"], g["form"],
                        g["qty"], " ".join(g["ids"])])

    print(f"locations={len(loc_id)} lots={len(lot_id)} drugs={len(drug_rows)} "
          f"units={len(groups)} skipped={skipped}")
    print(f"total quantity = {sum(g['qty'] for g in groups.values())}")
    print(f"wrote 01_inspect_legacy.sql, 02_delete_legacy.sql, 03_load_legacy.sql to {args.outdir}/")


if __name__ == "__main__":
    main()
