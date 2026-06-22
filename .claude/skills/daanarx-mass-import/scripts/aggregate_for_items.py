#!/usr/bin/env python3
"""
Parse a MASS export and aggregate for the core `items` schema: one record per
(form, specialty-bin, medication, dosage, unit) with quantity = count of
repeating containers. Medication name typos are corrected here. Output JSON is
fed to enrich_mass_items.mjs, which classifies + shapes it via the
@daana-health/domain-mass npm package.

Output (stdout or --out): JSON {records:[{bin, med, dosage, unit, sheet_form, quantity}]}
"""
import argparse, datetime, json, re, collections

SHEET_FORM = {"DATA-CARDS": "Card", "DATA-BOTTLES": "Bottle", "DATA-PSYCH": "Psych"}
SPEC_FIX = {"OCT": "OTC", "PSCYH": "PSYCH", "URO": "UROLOGY"}

# medication base-name typo corrections (case-insensitive, longest-first)
NAME_FIX = {
    "MDTHOCARBAMOL": "Methocarbamol", "METHOCARBAM": "Methocarbamol",
    "ESITALOPRAM": "Escitalopram", "ESTRESTO": "Entresto",
    "AMANTIDINE": "Amantadine", "ANASTRAZOLE": "Anastrozole",
    "BENZOTROPINE": "Benztropine", "CINCALCET": "Cinacalcet",
    "DIGOXINE": "Digoxin", "METROPROL": "Metoprolol",
    "MIDORINE": "Midodrine", "MIDODRONE": "Midodrine",
    "SPIRONOLATCTONE": "Spironolactone", "GLYCOPYRROLE": "Glycopyrrolate",
    "NYASTATIN": "Nystatin", "PROCHLORPER": "Prochlorperazine",
    "BUSPIRONE HYDROCHLORIDA": "Buspirone Hydrochloride",
    "DILTIAZAEM HYDROCHLORIDE": "Diltiazem Hydrochloride",
    "NITROGLYCERN": "Nitroglycerin", "LEVOTHYROXIN": "Levothyroxine",
    "AMITRIPTYLIN": "Amitriptyline", "AMYTRIPTYLIN": "Amitriptyline",
}


def clean(c):
    return "" if c is None else str(c).strip()


def norm_dose(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return f"{v.month}-{v.day}"
    s = clean(v)
    return s[:-2] if re.fullmatch(r"\d+\.0", s) else s


def canon_specialty(label):
    s = re.sub(r"\s+", " ", clean(label)).upper()
    return " ".join(SPEC_FIX.get(p, p) for p in s.split(" ")) if s else ""


def fix_name(med):
    m = re.sub(r"\s+", " ", clean(med)).strip()
    up = m.upper()
    for bad in sorted(NAME_FIX, key=len, reverse=True):
        if up == bad or up.startswith(bad + " "):
            return NAME_FIX[bad] + m[len(bad):]
    # Title-case ALLCAPS donor names; leave mixed-case as-is
    return m.title() if m.isupper() else m


def read_rows(path, form_arg):
    out = []
    if path.lower().endswith((".xlsx", ".xlsm")):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        for sheet, form in SHEET_FORM.items():
            if sheet in wb.sheetnames:
                for r in wb[sheet].iter_rows(values_only=True):
                    out.append((form, list(r) + [None] * 6))
    else:
        import csv
        with open(path, newline="") as f:
            for r in csv.reader(f):
                out.append((form_arg, list(r) + [None] * 6))
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("input")
    ap.add_argument("--form", default="Card")
    ap.add_argument("--out", default="-")
    args = ap.parse_args()

    groups = collections.OrderedDict()
    skipped = 0
    for form, row in read_rows(args.input, args.form):
        lot, spec, med, st, un, qid = [clean(x) for x in row[:5]] + [clean(row[5])]
        med = fix_name(med)
        dose = norm_dose(row[3])
        un = "" if un.upper() == "X" else un.lower()
        if un == "meg":
            un = "meq"
        if not med or not qid:
            skipped += 1
            continue
        binlabel = canon_specialty(spec)
        key = (form, binlabel, med, dose, un)
        g = groups.setdefault(key, dict(sheet_form=form, bin=binlabel, med=med,
                                        dosage=dose, unit=un, quantity=0))
        g["quantity"] += 1

    recs = list(groups.values())
    payload = {"source": args.input.split("/")[-1], "records": recs}
    txt = json.dumps(payload, indent=2)
    if args.out == "-":
        print(txt)
    else:
        open(args.out, "w").write(txt)
    import sys
    print(f"records={len(recs)} total_qty={sum(r['quantity'] for r in recs)} "
          f"bins={len({r['bin'] for r in recs})} skipped={skipped}", file=sys.stderr)


if __name__ == "__main__":
    main()
